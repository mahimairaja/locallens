"""Spreadsheet extractor for .xlsx, .xls, .csv, .tsv files."""

import csv
from pathlib import Path

from rich.console import Console

from locallens.extractors.base import LocalLensExtractor

console = Console()

try:
    import openpyxl

    _openpyxl_available = True
except ImportError:
    _openpyxl_available = False


class SpreadsheetExtractor(LocalLensExtractor):
    """Extract text from spreadsheet and CSV files."""

    def supported_extensions(self) -> list[str]:
        return [".xlsx", ".xls", ".csv", ".tsv"]

    def name(self) -> str:
        return "spreadsheet"

    def extract(self, file_path: Path) -> str:
        """Extract spreadsheet content as structured text."""
        ext = file_path.suffix.lower()
        if ext in (".csv", ".tsv"):
            return self._extract_csv(file_path, ext)
        elif ext in (".xlsx", ".xls"):
            return self._extract_xlsx(file_path)
        return ""

    def _extract_csv(self, file_path: Path, ext: str) -> str:
        """Extract CSV/TSV as key-value text lines."""
        try:
            delimiter = "\t" if ext == ".tsv" else ","
            with open(file_path, encoding="utf-8", errors="replace") as f:
                reader = csv.reader(f, delimiter=delimiter)
                rows = list(reader)

            if not rows:
                return ""

            headers = rows[0]
            lines = [f"Sheet: {file_path.stem}"]
            for row in rows[1:]:
                pairs = []
                for i, cell in enumerate(row):
                    col = headers[i] if i < len(headers) else f"col_{i}"
                    if cell.strip():
                        pairs.append(f"{col}: {cell}")
                if pairs:
                    lines.append("; ".join(pairs))

            return "\n".join(lines)
        except Exception as exc:
            console.print(
                f"[yellow]Warning: Could not extract CSV {file_path}: {exc}[/yellow]"
            )
            return ""

    def _extract_xlsx(self, file_path: Path) -> str:
        """Extract XLSX sheets as key-value text lines."""
        if not _openpyxl_available:
            console.print(
                f"[yellow]Warning: openpyxl not installed. Cannot extract {file_path}[/yellow]"
            )
            return ""
        try:
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            blocks = []
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows = list(ws.iter_rows(values_only=True))
                if not rows:
                    continue
                headers = [
                    str(h) if h is not None else f"col_{i}"
                    for i, h in enumerate(rows[0])
                ]
                lines = [f"Sheet: {sheet_name}"]
                for row in rows[1:]:
                    pairs = []
                    for i, cell in enumerate(row):
                        col = headers[i] if i < len(headers) else f"col_{i}"
                        val = str(cell) if cell is not None else ""
                        if val.strip():
                            pairs.append(f"{col}: {val}")
                    if pairs:
                        lines.append("; ".join(pairs))
                blocks.append("\n".join(lines))
            wb.close()
            return "\n\n".join(blocks)
        except Exception as exc:
            console.print(
                f"[yellow]Warning: Could not extract XLSX {file_path}: {exc}[/yellow]"
            )
            return ""
